# -*- coding:utf-8 -*-
import json
import os
import pickle

import xlwt


def read_with_pickle(name="D:/work/20190108_错别字/节目人工审核/数据/category_dict.pkl"):
    """
    读 pickle 文件
    :param name:
    :return:
    """
    with open(name, "rb") as f:
        category_dict = pickle.load(f)
    return category_dict


def dump_with_pickle(name="D:/work/20190108_错别字/节目人工审核/数据/category_dict.pkl", data={}):
    """
    读 pickle 文件
    :param data:
    :param name:
    :return:
    """
    with open(name, "wb") as f:
        pickle.dump(data, f)


def read_with_json(name='data.json'):
    """
    标签映射到单词
    :return:
    """
    with open(name, 'r', encoding='utf-8') as f:
        data = json.load(f)
        f.close()
    return data


def dump_with_json(name="data.json", data={}):
    # 写入 JSON 数据
    with open(name, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
        f.close()


def read_dict(f):
    """
    读入 dict
    文件格式：
        词,词频
    :param f:
    :return:
    """
    dic = {}
    with open(f, "r", encoding="utf-8") as f1:
        d = f1.readlines()
    for l in d:
        wc = l.split(",")
        if len(wc) == 2:
            dic[wc[0]] = int(wc[1])
    return dic


def dump_dict(f, d):
    """
    写入 dict
    {词频:数量}
    :param f:
    :param d:
    :return:
    """
    with open(f, 'w', encoding="utf-8") as fw:  # 写入 tf_idf 值
        for k, v in d.items():
            fw.write('%s,%d\n' % (k, v))


def read_items(f, split=","):
    """
    读入 文件

    [(词语,数量), (词语2,数量2)]
    :param split:
    :param f:
    :return:
    """
    items = []
    with open(f, "r", encoding="utf-8") as f1:
        d = f1.readlines()
    for l in d:
        wc = l.strip().split(split)
        if len(wc) == 2:
            items.append([wc[0], float(wc[1])])
    return items


def dump_items(f, items):
    """
    写入 items
    [(词语,数量), (词语2,数量2)]
    :param f:
    :param items:
    :return:
    """
    with open(f, 'w', encoding="utf-8") as fw:  # 写入 tf_idf 值
        for i in items:
            i = [str(b) for b in i]
            s = ",".join(i)
            fw.write('{}\n'.format(s))


def dump_list(f, ls, split=" "):
    """
    写入 items
    [[w11,w12,w13], [w21,w22,w23]]
    :param split:
    :param f:
    :param ls:
    :return:
    """
    with open(f, 'w', encoding="utf-8") as fw:  # 写入 tf_idf 值
        for l in ls:
            l = [str(i) for i in l]
            fw.write(split.join(l) + "\n")


def read_lines(file_name, encoding="utf-8"):
    """
    读入文本信息，返回 lines 会将 \n 一起读出来
    :param encoding:
    :param file_name:
    :return:
    """
    f = open(file_name, 'r', encoding=encoding)
    lines = f.readlines()
    f.close()
    return lines


def write_lines(file_name, lines):
    """
    写入文本信息
    :param file_name:
    :param lines: 每一行 line 后面添加 \n 这样写入的文本才会换行
    :return:
    """
    f = open(file_name, "w", encoding="utf-8")
    f.writelines(lines)
    f.close()


import re


def article_split(f):
    """
    将一个文件 拆分为 章节
    :param dir:  存放章节的目录
    :param f: 文件
    :return:
    """
    lines = read_lines(f)
    s = True
    p = 0
    a_count = 1
    for (l, index) in zip(lines, range(len(lines))):
        t = l.strip()  # 去除 左右空格
        if t == "------------":  # 开始分章节
            s = not s
        elif re.match(
                " [\u4e00-\u9fa5a-zA-Z0-9].*|\S.*|　　第[0-9零一二三四五六七八九十百千]+章 [\u4e00-\u9fa5a-zA-Z0-9]+|    <!--阅读面页章节尾部广告-->第[0-9一零二三四五六七八九十百千]+章 .*",
                l) is not None:  # 开始分章节
            s = not s
        if not s or (s and index == len(lines) - 1):  # 一章结束 或者 整个文本结束
            if index == len(lines) - 1:  # 如果是最后一行
                ls = lines[p + 1:]
            else:  # 如果不是最后一行
                ls = lines[p + 1:index]
            p = index
            s = True
            if len(ls) < 10 and sum([len(l) for l in ls]) < 100:  # 小于 10 行的文本，且总字数不足100字的，不为一个章节
                continue
            (filepath, tempfilename) = os.path.split(f)
            (filename, extension) = os.path.splitext(tempfilename)
            a_dir = os.path.join(filepath, filename)  # 建立一个目录 存放 章节内容
            if not os.path.exists(a_dir):
                os.mkdir(a_dir)
            # 写入这一章内容 章节名称：文件名+章节次数.txt
            write_lines(file_name=a_dir + "/" + filename + str(a_count) + extension, lines=ls)
            a_count += 1


def write_excel_xls(sheet_names, sheet_data, file):
    """
    写入 excel xls 格式

    注意：xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535
    :param sheet_names:
    :param sheet_data:{sheet_name:[[x,x], [x,x]]}
    :param file:
    :return:
    """
    f = xlwt.Workbook()
    for sheet_name in sheet_names:
        sheet = f.add_sheet(sheet_name, cell_overwrite_ok=True)
        data = sheet_data[sheet_name]
        for i in range(len(data)):
            for j in range(len(data[i])):
                sheet.write(i, j, data[i][j])
    f.save(file)


from openpyxl import Workbook


def write_excel_xlsx(sheet_names, sheet_data, file):
    """
    写入 excel xlsx 格式

    注意：xlsx 最大行数达到1048576
    :param sheet_names:
    :param sheet_data:{sheet_name:[[x,x], [x,x]]}
    :param file:
    :return:
    """
    outwb = Workbook()  # 打开一个将写的文件
    for sheet_name in sheet_names:
        outws = outwb.create_sheet(title=sheet_name)  # 在将写的文件创建sheet
        data = sheet_data[sheet_name]
        for i in range(len(data)):
            for j in range(len(data[i])):
                outws.cell(row=(i + 1), column=(j + 1)).value = data[i][j]  # 写文件
    outwb.save(file)  # 一定要记得保存


from openpyxl import load_workbook


def read_excel_xlsx(f, sheet_name):
    """
    读入 excel xlsx 格式文件

    :param f:
    :param sheet_name:
    :return:
    """
    # 默认可读写，若有需要可以指定write_only和read_only为True
    wb = load_workbook(f)
    # 根据sheet名字获得sheet
    sheet = wb.get_sheet_by_name(sheet_name)
    data = [[cell.value for cell in row] for row in sheet.rows]
    return data


def read_excel_xlsx_names(f, sheet_names):
    """
    读入 excel xlsx 格式文件

    :param f:
    :param sheet_names: sheet 表单列表
    :return:
    """
    sheet_data = {}
    # 默认可读写，若有需要可以指定write_only和read_only为True
    wb = load_workbook(f)
    # 根据sheet名字获得sheet
    for sheet_name in sheet_names:
        sheet = wb.get_sheet_by_name(sheet_name)
        data = [[cell.value for cell in row] for row in sheet.rows]
        sheet_data[sheet_name] = data
    return sheet_data


def rm_rf(file):
    """
    删除文件
    如果是
    :param file:
    :return:
    """
    pass


import requests


def download_f(url, f):
    r = requests.get(url)
    with open(f, "wb") as code:
        code.write(r.content)


def down_load_novel():
    path = "D:/work/20190408_关键词/"
    excel_f = "novel-path.xlsx"
    data = read_excel_xlsx(os.path.join(path, excel_f), "data")
    print(data[0:2])
    title_url = filter(lambda x: x[1].startswith("http://"), [(row[4], row[10]) for row in data])
    novel_dir = os.path.join(path, "4000本小说")
    count = 0
    n_set = set()
    for (title, url) in title_url:
        count += 1
        print(title)
        n_set.add(title)
        # download_f(url, os.path.join(novel_dir, title + ".zip"))
    print(count)
    print(len(n_set))


import zipfile


def extract_zip(o_f, t_dir):
    zFile = zipfile.ZipFile(o_f, "r")
    # ZipFile.namelist(): 获取ZIP文档内所有文件的名称列表
    for fileM in zFile.namelist():
        filename = fileM.encode('cp437').decode('gbk')
        if os.path.exists(os.path.join(t_dir, filename)):  # 如果已经有了对应文件，不解压这个文件
            continue
        zFile.extract(fileM, t_dir)
        os.rename(os.path.join(t_dir, fileM), os.path.join(t_dir, filename))
    zFile.close()


def batch_extract_zip():
    path = "D:/work/20190408_关键词/"
    novel_dir = os.path.join(path, "4000本小说")
    t_dir = os.path.join(path, "4000本小说-抽取")
    ns = os.listdir(novel_dir)
    for n in ns:
        print(n + "开始抽取")
        o_f = os.path.join(novel_dir, n)
        extract_zip(o_f, t_dir)


def count_file_lines(filename):
    """
    快速统计文件行数
    :param filename:
    :return:
    """
    count = 0
    fp = open(filename, "rb")
    byte_n = bytes("\n", encoding="utf-8")
    while 1:
        buffer = fp.read(16 * 1024 * 1024)
        if not buffer:
            count += 1  # 包含最后一行空行 ''
            break
        count += buffer.count(byte_n)
    fp.close()
    return count


def split_file(path, num=5):
    """
    拆分文件
    将一个大文件 拆分为多个小文件
    :param path:
    :param num:
    :return:
    """
    count = count_file_lines(path)
    print(count)
    temp = int((count + 1) / num)
    f = open(path, 'r', encoding="utf-8")  # 重新读入文件
    line = f.readline()
    count = 0
    file_count = 0
    while line:
        count += 1  # 记录当前文件输入行数
        if count == 1:  # 刚开始，此时创建文件
            file_count += 1  # 当前文件顺序
            f2 = open(path + str(file_count), "a", encoding="utf-8")  # 打开一个文件用于追加数据
        # 正常情况下，不停朝着文件追加数据
        f2.write(line)
        if count == temp:  # 当前文件输入到达指定行数，关闭当前文件，count置成0
            f2.close()
            count = 0
        line = f.readline()  # 读下一行
    if count > 0:  # 说明最后还有一波儿数据正在添加，但是不够 写一页 就已经写完了，此时关闭文件
        f2.close()
    f.close()  # 关闭总的文件


if __name__ == '__main__':
    pass
